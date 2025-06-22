from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_from_directory
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from database import teachers_collection, students_collection, classes_collection, attendance_collection
from bson.objectid import ObjectId
import os
from dotenv import load_dotenv
import uuid
from flask import Flask, render_template, request, redirect, session, url_for
from pymongo import MongoClient
# New import at the top
from encode_faces import register_face, delete_face_encoding
from database import attendance_collection
from datetime import datetime
import face_recognition
import numpy as np
from datetime import datetime
from collections import defaultdict
from bson import ObjectId
import pickle
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
import os.path

# Load environment variables
load_dotenv()

# Initialize database with admin user
from init_db import init_db
init_db()

# Create directory for Excel sheets
EXCEL_FOLDER = 'static/attendance_sheets'
os.makedirs(EXCEL_FOLDER, exist_ok=True)

def save_to_excel(class_id, records):
    """
    Save attendance records to Excel file organized by class
    """
    try:
        class_info = classes_collection.find_one({"_id": ObjectId(class_id)})
        if not class_info:
            print("Error: Class not found")
            return None
        
        # Create a valid filename (remove any invalid characters)
        class_name = "".join([c for c in class_info['name'] if c.isalnum() or c in (' ', '-', '_')])
        excel_path = os.path.join(EXCEL_FOLDER, f"{class_name}_attendance.xlsx")
        
        # Ensure the directory exists
        os.makedirs(os.path.dirname(excel_path), exist_ok=True)
        
        # Load existing workbook or create new one
        if os.path.exists(excel_path):
            wb = load_workbook(excel_path)
        else:
            wb = Workbook()
            # Set up header style
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
            
            # Create and style the header row
            ws = wb.active
            ws.title = class_name
            headers = ['Date', 'Time', 'Roll No', 'Name', 'Status', 'Source']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')

        ws = wb.active
        next_row = ws.max_row + 1

        # Add new records
        for record in records:
            row_data = [
                record['date'],
                record['time'],
                record['roll_no'],
                record['name'],
                record['status'],
                record.get('source', 'unknown')
            ]
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=next_row, column=col, value=value)
                cell.alignment = Alignment(horizontal='center')
                # Highlight present/absent status
                if col == 5:  # Status column
                    if value == 'Present':
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    else:
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            next_row += 1

        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            try:
                max_length = max(len(str(cell.value)) for cell in column if cell.value)
            except ValueError:
                max_length = 10
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width

        # Save the workbook
        wb.save(excel_path)
        print(f"Excel file saved successfully: {excel_path}")
        return excel_path
    except Exception as e:
        print(f"Error saving Excel file: {str(e)}")
        return None

app = Flask(__name__)
app.secret_key = os.getenv('SECRET_KEY', 'supersecret')  # Use environment variable for secret key

# Create directories if they don't exist
UPLOAD_FOLDER = 'static/student_photos'
EXCEL_FOLDER = 'static/attendance_sheets'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(EXCEL_FOLDER, exist_ok=True)
os.makedirs('encodings', exist_ok=True)
os.makedirs('static/class_photos', exist_ok=True)

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

login_manager = LoginManager()
login_manager.login_view = "login"
login_manager.init_app(app)

# Dummy User class using Flask-Login
class User(UserMixin):
    def __init__(self, user_id, username, email=None, created_at=None, photo_path=None):
        self.id = user_id
        self.username = username
        self.email = email
        self.created_at = created_at
        self.photo_path = photo_path

@login_manager.user_loader
def load_user(user_id):
    teacher = teachers_collection.find_one({"_id": ObjectId(user_id)})
    if teacher:
        return User(
            str(teacher["_id"]),
            teacher["username"],
            teacher.get("email"),
            teacher.get("created_at"),
            teacher.get("photo_path")
        )
    return None

# Route handlers
@app.route("/")
def home():
    if current_user.is_authenticated:
        return redirect("/dashboard")
    return render_template("landing.html")

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form["username"]
        password = request.form["password"]

        teacher = teachers_collection.find_one({"username": username, "password": password})
        if teacher:
            user_obj = User(
                str(teacher["_id"]),
                teacher["username"],
                teacher.get("email"),
                teacher.get("created_at"),
                teacher.get("photo_path")
            )
            login_user(user_obj)
            flash("Logged in successfully!")
            return redirect("/dashboard")
        else:
            flash("Invalid credentials")
            return redirect("/login")

    return render_template("login.html")

@app.route("/register", methods=["GET", "POST"])
def register():
    if request.method == "POST":
        username = request.form["username"]
        email = request.form.get("email")
        password = request.form["password"]
        confirm_password = request.form["confirm_password"]

        # Check if passwords match
        if password != confirm_password:
            flash("Passwords do not match!")
            return redirect(url_for("register"))

        # Check if username already exists
        if teachers_collection.find_one({"username": username}):
            flash("Username already exists!")
            return redirect(url_for("register"))

        # Create new teacher account
        teacher_data = {
            "username": username,
            "email": email,
            "password": password,  # In a production environment, you should hash the password
            "created_at": datetime(2025, 4, 30)
        }
        teachers_collection.insert_one(teacher_data)
        flash("Registration successful! Please login.")
        return redirect(url_for("login"))

    return render_template("register.html")

@app.route("/dashboard")
@login_required
def dashboard():
    # --- Analytics Stats ---
    total_students = students_collection.count_documents({})
    now = datetime.now()
    month_start = datetime(now.year, now.month, 1)
    month_end = datetime(now.year, now.month, 30)
    month_records = list(attendance_collection.find({
        "date": {"$gte": month_start.strftime('%Y-%m-%d'), "$lte": month_end.strftime('%Y-%m-%d')}
    }))
    month_present = sum(1 for r in month_records if r.get('status') == 'Present')
    month_total = len(month_records)
    attendance_percent = round((month_present / month_total) * 100, 1) if month_total else 0
    today_str = now.strftime('%Y-%m-%d')
    absent_today = attendance_collection.count_documents({"date": today_str, "status": "Absent"})
    image_records = list(attendance_collection.find({"source": "image"}))
    image_present = sum(1 for r in image_records if r.get('status') == 'Present')
    image_total = len(image_records)
    image_accuracy = round((image_present / image_total) * 100, 1) if image_total else 0
    # --- End Analytics Stats ---
    return render_template("dashboard.html", username=current_user.username,
        total_students=total_students,
        attendance_percent=attendance_percent,
        absent_today=absent_today,
        image_accuracy=image_accuracy
    )

@app.route("/logout")
@login_required
def logout():
    logout_user()
    flash("Logged out successfully.")
    return redirect("/login")

@app.route("/profile")
@login_required
def profile():
    # Get teacher info
    teacher = teachers_collection.find_one({"_id": ObjectId(current_user.id)})
    total_classes = classes_collection.count_documents({"created_by": current_user.id})
    total_students = students_collection.count_documents({"class_id": {"$in": [c["_id"] for c in classes_collection.find({"created_by": current_user.id})]}})
    # For demo, recent_activities is empty
    recent_activities = []
    # Patch current_user with missing fields if needed
    email = teacher.get("email", "-")
    created_at = teacher.get("created_at")
    if not created_at:
        created_at = datetime(2025, 4, 30)
    class PatchedUser:
        def __init__(self, user):
            self.id = user.id
            self.username = user.username
            self.email = email
            self.created_at = created_at
            self.photo_path = teacher.get("photo_path")
    patched_user = PatchedUser(current_user)
    return render_template(
        "profile.html",
        current_user=patched_user,
        total_classes=total_classes,
        total_students=total_students,
        recent_activities=recent_activities
    )

# Path to store student images
UPLOAD_FOLDER = 'static/student_photos'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

@app.route("/register_student", methods=["GET", "POST"])
@login_required
def register_student():
    if request.method == "POST":
        name = request.form["name"]
        roll_no = request.form["roll_no"]
        class_id = request.form["class_id"]
        photo = request.files["photo"]

        if photo:
            filename = f"{roll_no}_{photo.filename}"
            photo_path = f"student_photos/{filename}"  # Only relative to static/
            save_path = os.path.join("static", "student_photos", filename)
            os.makedirs(os.path.dirname(save_path), exist_ok=True)
            photo.save(save_path)

            student_data = {
                "name": name,
                "roll_no": roll_no,
                "class_id": ObjectId(class_id),
                "photo_path": photo_path
            }

            try:
                print("Attempting to insert:", student_data)
                result = students_collection.insert_one(student_data)
                print("Insert result:", result.inserted_id)

                try:
                    image = face_recognition.load_image_file(save_path)
                    encodings = face_recognition.face_encodings(image)

                    if not encodings:
                        raise ValueError("No face detected in the image.")
                    if len(encodings) > 1:
                        print(f"⚠️ Warning: Multiple faces detected in {photo_path}. Using the first one.")

                    face_encoding = encodings[0]

                    os.makedirs("encodings", exist_ok=True)
                    with open(os.path.join("encodings", f"{roll_no}.pkl"), "wb") as f:
                        pickle.dump({
                            "roll_number": roll_no,
                            "encoding": face_encoding
                        }, f)

                    flash("Student registered and face encoding saved!")
                except Exception as e:
                    flash(f"Student registered, but face encoding failed: {str(e)}", "danger")

            except Exception as e:
                print("Error while inserting student:", e)
                flash("Error while registering student.")
            return redirect("/dashboard")

    classes = list(classes_collection.find({"created_by": current_user.id}))
    return render_template("register_student.html", classes=classes)

@app.route('/view_students')
@login_required
def view_students():
    class_id = request.args.get('class_id')
    # Get all classes for the filter dropdown
    classes = list(classes_collection.find())
    pipeline = [
        {
            '$lookup': {
                'from': 'classes',
                'localField': 'class_id',
                'foreignField': '_id',
                'as': 'class_info'
            }
        },
        {
            '$addFields': {
                'class_name': {
                    '$ifNull': [
                        {'$arrayElemAt': ['$class_info.name', 0]},
                        'Not Assigned'
                    ]
                }
            }
        }
    ]
    if class_id:
        pipeline.insert(0, {'$match': {'class_id': ObjectId(class_id)}})
    students = list(students_collection.aggregate(pipeline))
    for student in students:
        student['_id'] = str(student['_id'])
        if 'class_id' in student:
            student['class_id'] = str(student['class_id'])
    # Calculate attendance percentage for each student in the selected class
    for student in students:
        total_attendance = attendance_collection.count_documents({
            "class_id": ObjectId(student["class_id"]) if "class_id" in student else None,
            "roll_no": student["roll_no"]
        })
        present_count = attendance_collection.count_documents({
            "class_id": ObjectId(student["class_id"]) if "class_id" in student else None,
            "roll_no": student["roll_no"],
            "status": "Present"
        })
        attendance_percent = round((present_count / total_attendance) * 100, 1) if total_attendance else 0
        student["attendance_percent"] = attendance_percent
    return render_template('view_students.html', students=students, classes=classes, selected_class=class_id)


@app.route('/delete_student/<student_id>', methods=['POST'])
@login_required
def delete_student(student_id):
    student = students_collection.find_one({'_id': ObjectId(student_id)})
    if student:
        photo_path = student['photo_path']
        roll_no = student['roll_no']  # Needed for encoding deletion

        # Delete photo
        if os.path.exists(photo_path):
            os.remove(photo_path)

        # Delete face encoding
        delete_face_encoding(roll_no)

        # Delete from DB
        students_collection.delete_one({'_id': ObjectId(student_id)})

    return redirect(url_for('view_students'))


@app.route('/register_all_faces', methods=['POST'])
@login_required
def register_all_faces():
    students = students_collection.find()
    success_count = 0
    fail_count = 0

    for student in students:
        roll_no = student.get("roll_no")
        photo_path = student.get("photo_path")

        if not roll_no or not photo_path or not os.path.exists(photo_path):
            fail_count += 1
            continue

        try:
            register_face(photo_path, roll_no)
            success_count += 1
        except Exception as e:
            print(f"Failed to register {roll_no}: {e}")
            fail_count += 1

    flash(f"Successfully registered encodings for {success_count} students. Failed: {fail_count}", "info")
    return redirect(url_for('dashboard'))


from datetime import datetime  # Make sure this import is at the top

# === Manual Attendance ===
@app.route('/mark_attendance_manual', methods=['GET', 'POST'])
@login_required
def mark_attendance_manual():
    selected_class = request.args.get('class_id', None)
    classes = list(classes_collection.find({"created_by": current_user.id}))
    
    if selected_class:
        students = list(students_collection.find({"class_id": ObjectId(selected_class)}))
    else:
        students = []

    if request.method == 'POST':
        marked = request.form.getlist('attendance')
        today = datetime.now().strftime('%Y-%m-%d')
        now_time = datetime.now().strftime('%H:%M:%S')
        records_to_insert = []
        for student in students:
            record = {
                'date': today,
                'time': now_time,
                'source': 'manual',
                'class_id': ObjectId(selected_class),
                'roll_no': student['roll_no'],
                'name': student['name'],
                'status': 'Present' if student['roll_no'] in marked else 'Absent'
            }
            records_to_insert.append(record)
        
        if records_to_insert:
            attendance_collection.insert_many(records_to_insert)

        # Save to Excel
        attendance_records = list(attendance_collection.find({
            'date': today,
            'time': now_time,
            'class_id': ObjectId(selected_class)
        }))
        excel_path = save_to_excel(selected_class, attendance_records)
        
        flash("Attendance marked manually and saved to Excel.")
        return redirect('/dashboard')
    
    return render_template('mark_attendance_manual.html', 
                         students=students, 
                         classes=classes, 
                         selected_class=selected_class)

@app.route('/mark_attendance_image', methods=['GET', 'POST'])
@login_required
def mark_attendance_image():
    selected_class = request.args.get('class_id', None)
    classes = list(classes_collection.find({"created_by": current_user.id}))

    if request.method == 'POST':
        class_photos = request.files.getlist('class_photos')
        selected_class = request.form.get('class_id')
        recognized_rolls = set()

        # Load known encodings for students in the selected class
        students_in_class = list(students_collection.find({"class_id": ObjectId(selected_class)}))
        known_encodings = []
        roll_numbers = []
        
        for student in students_in_class:
            roll_no = student['roll_no']
            encoding_path = os.path.join('encodings', f"{roll_no}.pkl")
            if os.path.exists(encoding_path):
                with open(encoding_path, 'rb') as f:
                    data = pickle.load(f)
                    known_encodings.append(data["encoding"])
                    roll_numbers.append(roll_no)

        # Process all uploaded photos
        for photo in class_photos:
            if photo:
                file_path = os.path.join('static/class_photos', photo.filename)
                os.makedirs(os.path.dirname(file_path), exist_ok=True)
                photo.save(file_path)

                img = face_recognition.load_image_file(file_path)
                img_encodings = face_recognition.face_encodings(img)

                for face_encoding in img_encodings:
                    matches = face_recognition.compare_faces(known_encodings, face_encoding, tolerance=0.5)
                    for i, match in enumerate(matches):
                        if match:
                            recognized_rolls.add(roll_numbers[i])
                            break

        # Mark attendance for students using batch insertion
        today = datetime.now().strftime('%Y-%m-%d')
        now_time = datetime.now().strftime('%H:%M:%S')
        
        records_to_insert = []
        for student in students_in_class:
            record = {
                'date': today,
                'time': now_time,
                'source': 'image',
                'class_id': ObjectId(selected_class),
                'roll_no': student['roll_no'],
                'name': student['name'],
                'status': 'Present' if student['roll_no'] in recognized_rolls else 'Absent'
            }
            records_to_insert.append(record)
        
        if records_to_insert:
            attendance_collection.insert_many(records_to_insert)

        # Save to Excel
        attendance_records = list(attendance_collection.find({
            'date': today,
            'time': now_time,
            'class_id': ObjectId(selected_class)
        }))
        excel_path = save_to_excel(selected_class, attendance_records)

        flash(f"Attendance marked for {len(recognized_rolls)} students and saved to Excel.")
        return redirect('/dashboard')

    return render_template('mark_attendance_image.html', 
                         classes=classes, 
                         selected_class=selected_class)

# View Attendance (Step 1 - Choose Type)
@app.route('/view_attendance', methods=['GET', 'POST'])
@login_required
def view_attendance():
    from datetime import datetime
    classes = list(classes_collection.find({"created_by": current_user.id}))
    # --- Analytics Stats ---
    total_students = students_collection.count_documents({})
    # Attendance % This Month
    now = datetime.now()
    month_start = datetime(now.year, now.month, 1)
    month_end = datetime(now.year, now.month, 30)  # April has 30 days
    month_records = list(attendance_collection.find({
        "date": {"$gte": month_start.strftime('%Y-%m-%d'), "$lte": month_end.strftime('%Y-%m-%d')}
    }))
    month_present = sum(1 for r in month_records if r.get('status') == 'Present')
    month_total = len(month_records)
    attendance_percent = round((month_present / month_total) * 100, 1) if month_total else 0
    # Absent Today
    today_str = now.strftime('%Y-%m-%d')
    absent_today = attendance_collection.count_documents({"date": today_str, "status": "Absent"})
    # Image Accuracy %
    image_records = list(attendance_collection.find({"source": "image"}))
    image_present = sum(1 for r in image_records if r.get('status') == 'Present')
    image_total = len(image_records)
    image_accuracy = round((image_present / image_total) * 100, 1) if image_total else 0
    # --- End Analytics Stats ---
    if request.method == 'POST':
        selected_class = request.form['class_id']
        return redirect(url_for('view_attendance_sessions_all', class_id=selected_class))
    return render_template('choose_attendance_type.html', classes=classes,
        total_students=total_students,
        attendance_percent=attendance_percent,
        absent_today=absent_today,
        image_accuracy=image_accuracy
    )

@app.route('/view_attendance/all/<class_id>')
@login_required
def view_attendance_all(class_id):
    records = list(attendance_collection.find({"class_id": ObjectId(class_id)}))
    class_info = classes_collection.find_one({"_id": ObjectId(class_id)})
    return render_template('view_attendance_session.html', 
                         records=records, 
                         class_name=class_info['name'] if class_info else None)

# View Attendance (Step 2 - List Sessions for that Type)
@app.route('/view_attendance/sessions/<source>/<class_id>')
@login_required
def view_attendance_sessions(source, class_id):
    pipeline = [
        {"$match": {
            "source": source,
            "class_id": ObjectId(class_id)
        }},
        {"$group": {"_id": {"date": "$date", "time": "$time"}}},
        {"$sort": {"_id.date": -1, "_id.time": -1}}
    ]
    sessions = list(attendance_collection.aggregate(pipeline))
    class_info = classes_collection.find_one({"_id": ObjectId(class_id)})
    return render_template('list_sessions.html', 
                         sessions=sessions, 
                         source=source, 
                         class_name=class_info['name'] if class_info else 'Unknown Class')

@app.route('/view_attendance/sessions/all/<class_id>')
@login_required
def view_attendance_sessions_all(class_id):
    pipeline = [
        {"$match": {"class_id": ObjectId(class_id)}},
        {"$group": {"_id": {"date": "$date", "time": "$time"}}},
        {"$sort": {"_id.date": -1, "_id.time": -1}}
    ]
    sessions = list(attendance_collection.aggregate(pipeline))
    class_info = classes_collection.find_one({"_id": ObjectId(class_id)})
    return render_template('list_sessions.html', 
                         sessions=sessions, 
                         class_name=class_info['name'] if class_info else 'Unknown Class',
                         class_id=class_id)

@app.route('/view_attendance/session/all/<date>/<time>/<class_id>')
@login_required
def view_attendance_session_all(date, time, class_id):
    query = {
        "date": date,
        "time": time,
        "class_id": ObjectId(class_id)
    }
    records = list(attendance_collection.find(query))
    class_info = classes_collection.find_one({"_id": ObjectId(class_id)})
    # --- Add photo_path to each attendance record ---
    student_photo_map = {s['roll_no']: s.get('photo_path') for s in students_collection.find({}, {'roll_no': 1, 'photo_path': 1})}
    for record in records:
        record['photo_path'] = student_photo_map.get(record['roll_no'], None)    # Get Excel file path
    class_name = class_info['name'] if class_info else ''
    # Sanitize class name for filename
    safe_class_name = "".join(c for c in class_name if c.isalnum() or c in (' ', '-', '_')).strip()
    excel_filename = f"{safe_class_name}_attendance.xlsx"
    # Use forward slashes for URL path
    excel_path = f"attendance_sheets/{excel_filename}"

    return render_template("view_attendance_session.html", 
                         records=records, 
                         date=date, 
                         time=time, 
                         class_name=class_name,
                         excel_path=excel_path)

# View Attendance (Step 3 - Session Attendance)
@app.route('/view_attendance/session/<source>/<date>/<time>')
@login_required
def view_attendance_session(source, date, time):
    class_id = request.args.get('class_id')
    query = {
        "source": source,
        "date": date,
        "time": time
    }
    if class_id:
        query["class_id"] = ObjectId(class_id)
    
    records = list(attendance_collection.find(query))
    class_info = None
    if class_id:
        class_info = classes_collection.find_one({"_id": ObjectId(class_id)})
    # --- Add photo_path to each attendance record ---
    student_photo_map = {s['roll_no']: s.get('photo_path') for s in students_collection.find({}, {'roll_no': 1, 'photo_path': 1})}
    for record in records:
        record['photo_path'] = student_photo_map.get(record['roll_no'], None)
    return render_template("view_attendance_session.html", 
                         records=records, 
                         date=date, 
                         time=time, 
                         source=source,
                         class_name=class_info['name'] if class_info else None)

@app.route("/manage_classes", methods=["GET", "POST"])
@login_required
def manage_classes():
    if request.method == "POST":
        class_name = request.form["class_name"]
        class_data = {
            "name": class_name,
            "created_by": current_user.id
        }
        classes_collection.insert_one(class_data)
        flash("Class created successfully!")
        return redirect(url_for('manage_classes'))
    
    classes = list(classes_collection.find({"created_by": current_user.id}))
    return render_template("manage_classes.html", classes=classes)

@app.route("/delete_class/<class_id>", methods=["POST"])
@login_required
def delete_class(class_id):
    # Update students to remove them from this class
    students_collection.update_many(
        {"class_id": ObjectId(class_id)},
        {"$unset": {"class_id": ""}}
    )
    # Delete the class
    classes_collection.delete_one({"_id": ObjectId(class_id)})
    flash("Class deleted successfully!")
    return redirect(url_for('manage_classes'))

@app.route("/change_password", methods=["GET", "POST"])
@login_required
def change_password():
    if request.method == "POST":
        current_password = request.form["current_password"]
        new_password = request.form["new_password"]
        confirm_password = request.form["confirm_password"]

        teacher = teachers_collection.find_one({"_id": ObjectId(current_user.id)})
        if not teacher or teacher.get("password") != current_password:
            flash("Current password is incorrect!", "danger")
            return redirect(url_for("change_password"))

        if new_password != confirm_password:
            flash("New passwords do not match!", "danger")
            return redirect(url_for("change_password"))

        if new_password == current_password:
            flash("New password must be different from current password!", "danger")
            return redirect(url_for("change_password"))

        if len(new_password) < 6 or not any(c.isdigit() for c in new_password) or not any(c in '!@#$%^&*' for c in new_password):
            flash("Password must be at least 6 characters, include a number and a special character!", "danger")
            return redirect(url_for("change_password"))

        teachers_collection.update_one({"_id": ObjectId(current_user.id)}, {"$set": {"password": new_password}})
        flash("Password updated successfully!", "success")
        return redirect(url_for("profile"))

    return render_template("change_password.html")

@app.route('/edit_profile', methods=['POST'])
@login_required
def edit_profile():
    username = request.form.get('username', '').strip()
    photo = request.files.get('profile_photo')
    teacher = teachers_collection.find_one({'_id': ObjectId(current_user.id)})

    # Check for username conflict
    if username and username != teacher['username']:
        if teachers_collection.find_one({'username': username, '_id': {'$ne': ObjectId(current_user.id)}}):
            flash('Username already taken!', 'danger')
            return redirect(url_for('profile'))

    update_data = {}
    if username and username != teacher['username']:
        update_data['username'] = username
    if photo and photo.filename:
        ext = os.path.splitext(photo.filename)[1].lower()
        if ext not in ['.jpg', '.jpeg', '.png']:
            flash('Invalid file type. Only JPG and PNG allowed.', 'danger')
            return redirect(url_for('profile'))
        if photo.content_length and photo.content_length > 2 * 1024 * 1024:
            flash('File too large. Max 2MB.', 'danger')
            return redirect(url_for('profile'))
        filename = f"{current_user.id}_profile{ext}"
        photo_path = f"student_photos/{filename}"
        save_path = os.path.join("static", "student_photos", filename)
        os.makedirs(os.path.dirname(save_path), exist_ok=True)
        photo.save(save_path)
        update_data['photo_path'] = photo_path
    if update_data:
        teachers_collection.update_one({'_id': ObjectId(current_user.id)}, {'$set': update_data})
        # Refresh current_user session
        updated_teacher = teachers_collection.find_one({'_id': ObjectId(current_user.id)})
        user_obj = User(
            str(updated_teacher['_id']),
            updated_teacher['username'],
            updated_teacher.get('email'),
            updated_teacher.get('created_at'),
            updated_teacher.get('photo_path')
        )
        login_user(user_obj)
        flash('Profile updated successfully!', 'success')
    else:
        flash('No changes made.', 'info')
    return redirect(url_for('profile'))

@app.route('/delete_attendance/<record_id>', methods=['POST'])
@login_required
def delete_attendance(record_id):
    from bson import ObjectId
    attendance_collection.delete_one({'_id': ObjectId(record_id)})
    flash('Attendance record deleted.', 'success')
    # Redirect back to the referring page
    return redirect(request.referrer or url_for('view_attendance'))

@app.route('/delete_attendance_session/<class_id>/<date>/<time>', methods=['POST'])
@login_required
def delete_attendance_session(class_id, date, time):
    attendance_collection.delete_many({
        'class_id': ObjectId(class_id),
        'date': date,
        'time': time
    })
    flash('Attendance session deleted.', 'success')
    return redirect(request.referrer or url_for('view_attendance'))

@app.route('/live_attendance', methods=['GET'])
@login_required
def live_attendance():
    selected_class = request.args.get('class_id', None)
    classes = list(classes_collection.find({"created_by": current_user.id}))
    return render_template('live_attendance.html', 
                         classes=classes, 
                         selected_class=selected_class)

@app.route('/process_frame', methods=['POST'])
@login_required
def process_frame():
    if 'frame' not in request.files:
        return jsonify({'error': 'No frame provided'}), 400

    class_id = request.form.get('class_id')
    if not class_id:
        return jsonify({'error': 'No class selected'}), 400

    frame = request.files['frame']
    frame_path = os.path.join('static/class_photos', 'temp_frame.jpg')
    frame.save(frame_path)

    # Load known encodings for students in the selected class
    students_in_class = list(students_collection.find({"class_id": ObjectId(class_id)}))
    known_encodings = []
    roll_numbers = []
    student_info = {}
    
    for student in students_in_class:
        roll_no = student['roll_no']
        encoding_path = os.path.join('encodings', f"{roll_no}.pkl")
        if os.path.exists(encoding_path):
            with open(encoding_path, 'rb') as f:
                data = pickle.load(f)
                known_encodings.append(data["encoding"])
                roll_numbers.append(roll_no)
                student_info[roll_no] = {
                    'name': student['name'],
                    'roll_no': roll_no,
                    'photo_path': student.get('photo_path', '')
                }

    # Process the frame for face recognition
    img = face_recognition.load_image_file(frame_path)
    face_locations = face_recognition.face_locations(img)
    face_encodings = face_recognition.face_encodings(img, face_locations)

    recognized_students = []
    for face_encoding in face_encodings:
        matches = face_recognition.compare_faces(known_encodings, face_encoding, tolerance=0.5)
        for idx, match in enumerate(matches):
            if match:
                roll_no = roll_numbers[idx]
                student = student_info.get(roll_no)
                if student and student not in recognized_students:
                    recognized_students.append(student)

    # Clean up temp file
    try:
        os.remove(frame_path)
    except:
        pass

    return jsonify({'recognized_students': recognized_students})

@app.route('/save_live_attendance', methods=['POST'])
@login_required
def save_live_attendance():
    data = request.json
    class_id = data.get('class_id')
    recognized_students = data.get('recognized_students', [])

    if not class_id:
        return jsonify({'error': 'No class selected'}), 400

    # Get all students in the class
    students_in_class = list(students_collection.find({"class_id": ObjectId(class_id)}))
    
    # Current timestamp
    today = datetime.now().strftime('%Y-%m-%d')
    now_time = datetime.now().strftime('%H:%M:%S')

    # Mark attendance for all students
    records_to_insert = []
    recognized_roll_numbers = [student['roll_no'] for student in recognized_students]
    
    for student in students_in_class:
        record = {
            'date': today,
            'time': now_time,
            'source': 'live',
            'class_id': ObjectId(class_id),
            'roll_no': student['roll_no'],
            'name': student['name'],
            'status': 'Present' if student['roll_no'] in recognized_roll_numbers else 'Absent'
        }
        records_to_insert.append(record)
    
    if records_to_insert:
        attendance_collection.insert_many(records_to_insert)
        # Save to Excel
        excel_path = save_to_excel(class_id, records_to_insert)
        
    return jsonify({
        'success': True,
        'message': f'Attendance marked for {len(recognized_students)} students',
        'date': today,
        'time': now_time
    })

@app.route('/download_excel/<class_name>')
@login_required
def download_excel(class_name):
    safe_class_name = "".join(c for c in class_name if c.isalnum() or c in (' ', '-', '_')).strip()
    excel_filename = f"{safe_class_name}_attendance.xlsx"
    excel_path = os.path.join('static', 'attendance_sheets', excel_filename)
    
    if not os.path.exists(excel_path):
        flash('Excel file not found', 'error')
        return redirect(url_for('view_attendance'))
        
    return send_from_directory(
        'static/attendance_sheets',
        excel_filename,
        as_attachment=True,
        download_name=excel_filename
    )

if __name__ == "__main__":
    app.run(debug=True)
